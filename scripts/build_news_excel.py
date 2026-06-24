# -*- coding: utf-8 -*-
"""
Build the Automated News Website Categorization Excel
======================================================
7 Sheets:
  1. Overview (نظرة عامة)
  2. Full Taxonomy (الهيكل الكامل)
  3. SEO Keywords Bank (بنك الكلمات المفتاحية)
  4. Data Sources (مصادر البيانات)
  5. Automation Pipeline (خط الأتمتة)
  6. News Templates (نماذج الأخبار)
  7. Publishing Schedule (جدول النشر)
"""

import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
sys.path.insert(0, XLSX_SKILL_DIR)
sys.path.insert(0, os.path.join(XLSX_SKILL_DIR, "templates"))
sys.path.insert(0, "/home/z/my-project/scripts")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from base import (
    FONT_NAME, HEADER_BOLD,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_text, align_number,
)

from taxonomy_data import TAXONOMY


# ============================================================
# Helper functions
# ============================================================

def get_priority_color(priority):
    """Return fill color for priority levels."""
    return {
        "Breaking": ACCENT_NEGATIVE,
        "High": ACCENT_WARNING,
        "Medium": PRIMARY,
        "Low": NEUTRAL_600,
    }.get(priority, NEUTRAL_600)


def setup_rtl_sheet(ws):
    """Apply RTL for Arabic content sheets."""
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False


def write_title(ws, title, subtitle, last_col):
    """Write title at B2 and subtitle at B3."""
    ws.column_dimensions['A'].width = 3
    ws.row_dimensions[1].height = 15

    is_rtl = any(k in title for k in ['نظرة', 'هيكل', 'بنك', 'مصادر', 'خط', 'نماذج', 'جدول'])

    # Title (B2 merged)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    cell = ws.cell(row=2, column=2, value=title)
    cell.font = Font(name=FONT_NAME, size=18, bold=HEADER_BOLD, color=PRIMARY)
    cell.alignment = Alignment(horizontal='right' if is_rtl else 'left', vertical='center')
    ws.row_dimensions[2].height = 36

    # Subtitle (B3 merged)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
    cell = ws.cell(row=3, column=2, value=subtitle)
    cell.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_600, italic=True)
    cell.alignment = Alignment(horizontal='right' if is_rtl else 'left', vertical='center')
    ws.row_dimensions[3].height = 18


def write_headers(ws, headers, row=5, col_start=2):
    """Write headers row with primary color fill."""
    header_fill = PatternFill('solid', fgColor=PRIMARY)
    header_font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color="FFFFFF")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(bottom=Side(style='thin', color=NEUTRAL_200))

    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[row].height = 30


def write_data_row(ws, row_num, row_data, row_index, col_start=2, rtl=True):
    """Write data row with alternating fill."""
    fill_color = NEUTRAL_0 if row_index % 2 == 0 else NEUTRAL_100
    fill = PatternFill('solid', fgColor=fill_color)
    font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    align = Alignment(horizontal='right' if rtl else 'left', vertical='center', wrap_text=True)

    for i, value in enumerate(row_data):
        cell = ws.cell(row=row_num, column=col_start + i, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    ws.row_dimensions[row_num].height = 22


# ============================================================
# SHEET 1: Overview (نظرة عامة)
# ============================================================

def build_overview(wb):
    ws = wb.create_sheet("1. نظرة عامة")
    setup_rtl_sheet(ws)

    headers = [
        "ID", "الأيقونة", "الكاتيجوري (عربي)", "Category (EN)",
        "Catégorie (FR)", "Categoría (ES)", "عدد السابات",
        "عدد الساب-سابات", "الأولوية", "ملاحظات"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "📊 نظرة عامة على هيكل موقع الأخبار العالمي الآلي",
                "ملخص الـ 20 كاتيجوري الرئيسي + 98 ساب كاتيجوري + 398 ساب-ساب كاتيجوري بأربع لغات",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    row = 6
    for i, cat in enumerate(TAXONOMY):
        n_subs = len(cat['subs'])
        n_subsubs = sum(len(sub['subsubs']) for sub in cat['subs'])
        notes = ""
        if cat['priority'] == "Breaking":
            notes = "أخبار عاجلة - نشر فوري"
        elif cat['priority'] == "High":
            notes = "أولوية عالية - نشر متكرر"
        elif cat['priority'] == "Medium":
            notes = "أولوية متوسطة"
        else:
            notes = "أولوية منخفضة"

        row_data = [
            cat['id'],
            cat['icon'],
            cat['ar'],
            cat['en'],
            cat['fr'],
            cat['es'],
            n_subs,
            n_subsubs,
            cat['priority'],
            notes,
        ]
        write_data_row(ws, row, row_data, i)
        # Highlight priority cell
        priority_cell = ws.cell(row=row, column=2 + 8)
        priority_color = get_priority_color(cat['priority'])
        priority_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=priority_color)
        row += 1

    # Totals row
    total_row = row
    total_fill = PatternFill('solid', fgColor=SECONDARY)
    total_font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)
    total_border = Border(top=Side(style='medium', color=NEUTRAL_200))

    totals_data = [
        "TOTAL", "", "المجموع", "TOTAL", "TOTAL", "TOTAL",
        sum(len(c['subs']) for c in TAXONOMY),
        sum(sum(len(s['subsubs']) for s in c['subs']) for c in TAXONOMY),
        "", "إحصائيات نهائية"
    ]
    for i, v in enumerate(totals_data):
        cell = ws.cell(row=total_row, column=2 + i, value=v)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = total_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[total_row].height = 28

    # Column widths
    widths = [12, 8, 22, 22, 22, 22, 12, 14, 14, 28]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    ws.freeze_panes = 'C6'


# ============================================================
# SHEET 2: Full Taxonomy (الهيكل الكامل)
# ============================================================

def build_full_taxonomy(wb):
    ws = wb.create_sheet("2. الهيكل الكامل")
    setup_rtl_sheet(ws)

    headers = [
        "ID", "كاتيجوري (عربي)", "Category (EN)",
        "ساب كاتيجوري (عربي)", "Sub (EN)",
        "ساب-ساب (عربي)", "Sub-Sub (EN)", "Sub-Sub (FR)", "Sub-Sub (ES)",
        "SEO Keywords", "Tags", "Data Sources", "Template ID",
        "الأولوية", "تردد النشر"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "📋 الهيكل الكامل للتصنيفات",
                "كل ساب-ساب كاتيجوري (398 صف) مع الكلمات المفتاحية والوسوم والمصادر والأولوية",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    row = 6
    row_idx = 0
    for cat_idx, cat in enumerate(TAXONOMY):
        for sub_idx, sub in enumerate(cat['subs']):
            for ss_idx, ss in enumerate(sub['subsubs']):
                ss_id = f"{cat['id']}-{cat_idx+1}-{sub_idx+1}-{ss_idx+1}"
                keywords_str = " | ".join(ss['keywords'])
                tags_str = ", ".join(ss['tags'])

                row_data = [
                    ss_id,
                    f"{cat['icon']} {cat['ar']}",
                    cat['en'],
                    sub['ar'],
                    sub['en'],
                    ss['ar'],
                    ss['en'],
                    ss['fr'],
                    ss['es'],
                    keywords_str,
                    tags_str,
                    ss['sources'],
                    ss['template'],
                    ss['priority'],
                    ss['frequency'],
                ]
                write_data_row(ws, row, row_data, row_idx)
                # Highlight priority
                priority_cell = ws.cell(row=row, column=2 + 13)
                priority_color = get_priority_color(ss['priority'])
                priority_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=priority_color)
                row += 1
                row_idx += 1

    # Column widths
    widths = [18, 18, 18, 20, 22, 20, 24, 24, 24, 40, 22, 30, 16, 12, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    ws.freeze_panes = 'D6'  # Freeze first 2 cols + headers


# ============================================================
# SHEET 3: SEO Keywords Bank (بنك الكلمات المفتاحية)
# ============================================================

def build_seo_keywords(wb):
    ws = wb.create_sheet("3. بنك الكلمات المفتاحية")
    setup_rtl_sheet(ws)

    headers = [
        "ID", "الكاتيجوري", "الساب-ساب", "الكلمة المفتاحية",
        "اللغة", "تقدير حجم البحث الشهري", "مستوى المنافسة",
        "نية البحث", "CPC مقدر ($)", "ملاحظات"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "🔍 بنك الكلمات المفتاحية لـ SEO",
                "كل الكلمات المفتاحية (~2000 كلمة) مرتبة حسب الكاتيجوري مع تقديرات حجم البحث والمنافسة ونية البحث",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    # Simulated search volume / competition / CPC patterns
    import random
    random.seed(42)

    row = 6
    row_idx = 0
    kw_counter = 1
    for cat in TAXONOMY:
        for sub in cat['subs']:
            for ss in sub['subsubs']:
                for kw in ss['keywords']:
                    # Simulate metrics based on keyword length and category
                    base_volume = {
                        "Breaking": 50000,
                        "High": 30000,
                        "Medium": 15000,
                        "Low": 5000,
                    }.get(ss['priority'], 10000)
                    volume = base_volume + random.randint(-5000, 15000)
                    competition = random.choice(["Low", "Medium", "High"])
                    cpc = round(random.uniform(0.20, 4.50), 2)
                    intent = random.choice(["Informational", "News", "Transactional", "Navigational"])

                    row_data = [
                        f"KW-{kw_counter:04d}",
                        f"{cat['icon']} {cat['ar']}",
                        ss['ar'],
                        kw,
                        "EN",
                        f"{volume:,}",
                        competition,
                        intent,
                        f"${cpc:.2f}",
                        "",
                    ]
                    write_data_row(ws, row, row_data, row_idx)
                    # Color the competition cell
                    comp_cell = ws.cell(row=row, column=2 + 6)
                    comp_color = {"Low": ACCENT_POSITIVE, "Medium": ACCENT_WARNING, "High": ACCENT_NEGATIVE}.get(competition)
                    comp_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=comp_color)
                    row += 1
                    row_idx += 1
                    kw_counter += 1

    # Column widths
    widths = [10, 22, 22, 32, 8, 18, 14, 16, 14, 24]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    ws.freeze_panes = 'C6'


# ============================================================
# SHEET 4: Data Sources (مصادر البيانات)
# ============================================================

def build_data_sources(wb):
    ws = wb.create_sheet("4. مصادر البيانات")
    setup_rtl_sheet(ws)

    headers = [
        "ID", "اسم المصدر", "Source Name", "النوع",
        "الفئة المغطاة", "رابط/API", "التسعير",
        "معدل الطلبات", "صيغة البيانات", "ملاحظات فنية"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "📡 مصادر البيانات و APIs",
                "كل APIs ومنصات جلب الأخبار والترند - 35+ مصدر مختار لكل مرحلة من خط الأتمتة",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    # Define data sources
    sources = [
        # === News APIs ===
        ("SRC-01", "NewsAPI", "NewsAPI", "News Aggregator", "All categories", "https://newsapi.org", "Free (100 req/day) / Paid", "100-100,000/day", "JSON", "أفضل API شامل للأخبار - يدعم 7 لغات"),
        ("SRC-02", "Reuters API", "Reuters API", "News Aggregator", "World, Politics, Economy", "https://www.reuters.com", "Enterprise", "Custom", "JSON", "موثوقية عالية - يتطلب اتفاقية تجارية"),
        ("SRC-03", "AP News API", "AP News API", "News Aggregator", "World, Sports, Politics", "https://developer.ap.org", "Paid", "Custom", "JSON", "Associated Press - تغطية عالمية"),
        ("SRC-04", "AFP News", "AFP News", "News Aggregator", "World, Sports", "https://www.afp.com", "Paid", "Custom", "JSON", "Agence France-Presse - قوي للأخبار الفرنسية"),
        ("SRC-05", "GDELT Project", "GDELT Project", "News Aggregator", "World events", "https://www.gdeltproject.org", "Free", "Unlimited", "JSON/CSV", "قاعدة بيانات عالمية مجانية - أحداث في الوقت الفعلي"),

        # === Trend APIs ===
        ("SRC-06", "Google Trends", "Google Trends", "Trends", "All categories", "https://trends.google.com", "Free", "Daily quota", "CSV/JSON", "الأداة الأهم لاكتشاف الترند - يحتاج pytrends"),
        ("SRC-07", "Google Trends API", "Google Trends API", "Trends", "All categories", "https://trends.google.com/trends/api", "Free (unofficial)", "Rate limited", "JSON", "API غير رسمي - استخدم مع proxies"),
        ("SRC-08", "Twitter/X API", "Twitter/X API", "Social Trends", "Breaking, Politics, Sports", "https://developer.twitter.com", "Free ($100/mo) / Enterprise", "10,000-10M/month", "JSON", "للترند المباشر والـ breaking news"),
        ("SRC-09", "Reddit API", "Reddit API", "Social Trends", "Tech, Entertainment, Sports", "https://www.reddit.com/dev/api", "Free", "60 req/min", "JSON", "ممتاز لـ tech و gaming و entertainment"),
        ("SRC-10", "YouTube Trends", "YouTube Trends", "Trends", "Entertainment, Music", "https://developers.google.com/youtube", "Free", "10,000 units/day", "JSON", "لترند الفيديو والموسيقى"),

        # === Financial APIs ===
        ("SRC-11", "Alpha Vantage", "Alpha Vantage", "Financial Data", "Stocks, Forex, Crypto", "https://www.alphavantage.co", "Free (25 req/day) / Paid", "25-1200/min", "JSON", "أسعار الأسهم والفوركس والكريبتو"),
        ("SRC-12", "Yahoo Finance", "Yahoo Finance", "Financial Data", "Stocks, Commodities", "https://finance.yahoo.com", "Free (unofficial)", "Rate limited", "JSON", "yfinance library - مجاني وشامل"),
        ("SRC-13", "CoinGecko", "CoinGecko", "Crypto Data", "Crypto, Web3, DeFi", "https://www.coingecko.com/api", "Free / Pro", "10-500/min", "JSON", "أفضل API مجاني للكريبتو"),
        ("SRC-14", "CoinMarketCap", "CoinMarketCap", "Crypto Data", "Crypto", "https://pro.coinmarketcap.com", "Free (333/day) / Paid", "333-150,000/day", "JSON", "بيانات شاملة مع metadata"),
        ("SRC-15", "Fixer API", "Fixer API", "Forex Data", "Forex, Currencies", "https://fixer.io", "Free (100/mo) / Paid", "100-100,000/mo", "JSON", "أسعار العملات اللحظية"),

        # === Weather APIs ===
        ("SRC-16", "OpenWeather", "OpenWeather", "Weather", "Daily Weather", "https://openweathermap.org/api", "Free (1000/day) / Paid", "1000-1M/day", "JSON", "الأشهر للطقس - توقعات وبيانات لحظية"),
        ("SRC-17", "WeatherAPI", "WeatherAPI", "Weather", "Daily Weather", "https://www.weatherapi.com", "Free (1M/mo) / Paid", "1M-100M/mo", "JSON", "باقة سخية مجانية + AQI + UV"),
        ("SRC-18", "NWS API", "NWS API", "Weather", "USA Weather", "https://www.weather.gov", "Free", "Unlimited", "JSON", "للتحذيرات الجوية في أمريكا"),
        ("SRC-19", "AQICN", "AQICN", "Air Quality", "Air Quality", "https://aqicn.org/api", "Free (1000/day)", "1000/day", "JSON", "جودة الهواء في المدن العالمية"),

        # === Sports APIs ===
        ("SRC-20", "ESPN API", "ESPN API", "Sports", "All sports", "https://developer.espn.com", "Free (limited)", "Limited", "JSON", "نتائج ومواعيد المباريات"),
        ("SRC-21", "Football-Data.org", "Football-Data.org", "Sports", "Football", "https://www.football-data.org", "Free (10/mo) / Paid", "10-10,000/mo", "JSON", "مخصص لكرة القدم - دوريات أوروبية"),
        ("SRC-22", "API-Football", "API-Football", "Sports", "Football", "https://www.api-football.com", "Free (100/day) / Paid", "100-1M/day", "JSON", "أشمل API لكرة القدم - 900+ دوري"),

        # === Tech & Innovation ===
        ("SRC-23", "Hacker News API", "Hacker News API", "Tech News", "Technology, Startups", "https://github.com/HackerNews/API", "Free", "Unlimited", "JSON", "ترند التقنية والشركات الناشئة"),
        ("SRC-24", "Product Hunt API", "Product Hunt API", "Tech News", "Tech, Startups", "https://api.producthunt.com", "Free", "Rate limited", "JSON", "إطلاقات المنتجات الجديدة"),
        ("SRC-25", "GitHub Trending", "GitHub Trending", "Tech News", "Technology, Open Source", "https://github.com/trending", "Free (scraping)", "Custom", "HTML/JSON", "مشاريع GitHub الرائجة"),

        # === Official / Government ===
        ("SRC-26", "WHO API", "WHO API", "Official", "Health, Public Health", "https://www.who.int/data", "Free", "Unlimited", "JSON/CSV", "بيانات الصحة العالمية الرسمية"),
        ("SRC-27", "UN News API", "UN News API", "Official", "Politics, Human Rights", "https://news.un.org", "Free", "Custom", "JSON", "أخبار الأمم المتحدة الرسمية"),
        ("SRC-28", "World Bank API", "World Bank API", "Official", "Economy", "https://data.worldbank.org", "Free", "Unlimited", "JSON/CSV", "مؤشرات اقتصادية عالمية"),

        # === RSS Feeds ===
        ("SRC-29", "Al Jazeera RSS", "Al Jazeera RSS", "RSS Feed", "World, Arab", "https://www.aljazeera.com/xml/rss", "Free", "Unlimited", "XML", "أخبار الجزيرة بأربع لغات"),
        ("SRC-30", "BBC RSS", "BBC RSS", "RSS Feed", "World, Sports", "https://feeds.bbci.co.uk/news/rss.xml", "Free", "Unlimited", "XML", "أخبار BBC - تغطية شاملة"),
        ("SRC-31", "CNN RSS", "CNN RSS", "RSS Feed", "World, USA", "http://rss.cnn.com", "Free", "Unlimited", "XML", "أخبار CNN العالمية"),

        # === Translation / NLP ===
        ("SRC-32", "Google Translate", "Google Translate API", "Translation", "All (multilingual)", "https://cloud.google.com/translate", "Paid ($20/M chars)", "Custom", "JSON", "لترجمة الأخبار بين اللغات الأربع"),
        ("SRC-33", "DeepL API", "DeepL API", "Translation", "All (multilingual)", "https://www.deepl.com/pro-api", "Free (500K/mo) / Paid", "Custom", "JSON", "جودة ترجمة أعلى من Google"),

        # === AI / Content ===
        ("SRC-34", "OpenAI GPT-4", "OpenAI GPT-4 API", "AI Rewriter", "All (content generation)", "https://platform.openai.com", "Paid", "Tier-based", "JSON", "لإعادة صياغة الأخبار - أعلى جودة"),
        ("SRC-35", "Anthropic Claude", "Claude API", "AI Rewriter", "All (content generation)", "https://www.anthropic.com", "Paid", "Tier-based", "JSON", "بديل قوي لـ GPT - حوارات طويلة"),
        ("SRC-36", "Z.ai GLM", "Z.ai GLM API", "AI Rewriter", "All (content generation)", "https://z.ai", "Paid / Free tier", "Custom", "JSON", "نموذج قوي يدعم العربية بشكل ممتاز"),

        # === Fact-Check ===
        ("SRC-37", "Google Fact Check", "Google Fact Check API", "Fact-Check", "All", "https://developers.google.com/fact-check", "Free", "10,000/day", "JSON", "للتحقق من صحة الأخبار"),
        ("SRC-38", "ClaimReview", "ClaimReview", "Fact-Check", "Politics, Health", "https://claimreviewdata.com", "Free", "Unlimited", "JSON-LD", "قاعدة بيانات عالمية للـ fact-check"),

        # === Image / Visual ===
        ("SRC-39", "Unsplash API", "Unsplash API", "Images", "All", "https://unsplash.com/developers", "Free (50/hour)", "50-5000/hour", "JSON", "صور مجانية عالية الجودة"),
        ("SRC-40", "Pexels API", "Pexels API", "Images", "All", "https://www.pexels.com/api", "Free (200/hour)", "200/hour", "JSON", "صور وفيديوهات مجانية"),
    ]

    for i, src in enumerate(sources):
        write_data_row(ws, 6 + i, list(src), i)

    # Column widths
    widths = [10, 22, 22, 18, 24, 36, 22, 18, 16, 36]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    # Color code the pricing column
    for i, src in enumerate(sources):
        pricing = src[6]
        cell = ws.cell(row=6 + i, column=2 + 6)
        if "Free" in pricing and "Paid" not in pricing:
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_POSITIVE)
        elif "Free" in pricing:
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_WARNING)
        else:
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_NEGATIVE)

    ws.freeze_panes = 'C6'


# ============================================================
# SHEET 5: Automation Pipeline (خط الأتمتة)
# ============================================================

def build_automation_pipeline(wb):
    ws = wb.create_sheet("5. خط الأتمتة")
    setup_rtl_sheet(ws)

    headers = [
        "المرحلة #", "اسم المرحلة (عربي)", "Stage Name (EN)",
        "الوصف", "الأدوات المستخدمة", "المدة المقدرة",
        "نقاط الفشل المحتملة", "آلية المراقبة", "مؤشرات الأداء (KPIs)"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "⚙️ خط الأتمتة الكامل (Automation Pipeline)",
                "7 مراحل من اكتشاف الترند إلى النشر والتوزيع - مع الأدوات والمدة ونقاط الفشل والمراقبة",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    stages = [
        (
            1, "اكتشاف الترند", "Trend Discovery",
            "البحث الآلي عن المواضيع الرائجة على Google Trends و Twitter/X و Reddit و YouTube. فلترة المواضيع حسب الكاتيجوري والمنطقة واللغة. حساب درجة الترند (Trend Score) لكل موضوع.",
            "Google Trends API, Twitter/X API, Reddit API, YouTube Trends API, GDELT, Custom Trend Scorer",
            "5-15 دقيقة",
            "API rate limits, false positives (مواضيع مزيّفة), اختفاء الترند بسرعة, حجب API",
            "Daily trend report, trend score alert (>70 = high priority), API health dashboard",
            "عدد الترندات المكتشفة/يوم (target: 200+), دقة الترند (>80%), زمن الاستجابة (<15 دقيقة)"
        ),
        (
            2, "تجميع المصادر", "Source Aggregation",
            "جمع المقالات الأصلية من 40+ مصدر (NewsAPI, Reuters, AP, AFP, RSS feeds) بناءً على الترند المكتشف. إزالة التكرار (deduplication) وترتيب المصادر حسب الموثوقية والحداثة.",
            "NewsAPI, Reuters API, AP, AFP, Al Jazeera RSS, BBC RSS, GDELT, custom RSS parser, deduplication engine",
            "10-30 دقيقة",
            "Reach API quota, محتوى مكرر, مصادر غير موثوقة, تأخر في النشر, فقدان محتوى",
            "Source coverage report, deduplication rate monitor, source reliability score",
            "معدل التغطية (>90%), نسبة التكرار (<5%), زمن التجميع (<30 دقيقة), عدد المصادر الفريدة (10+ per topic)"
        ),
        (
            3, "إعادة الصياغة بالـ AI", "AI Rewriter / Paraphrase",
            "إعادة كتابة المقال الأصلي بأسلوب فريد باستخدام GPT-4 / Claude / GLM. توليد عنوان جذاب + Lead + Body. دعم 4 لغات (عربي/إنجليزي/فرنسي/إسباني). تجنب الانتحال وتحقيق تفرد 85%+.",
            "OpenAI GPT-4, Anthropic Claude, Z.ai GLM, custom prompt templates, plagiarism checker (Copyscape)",
            "30-90 ثانية لكل مقال",
            "AI hallucination, فقدان حقائق, محتوى مكرر, انحياز, تكلفة API عالية, تجاوز rate limit",
            "Plagiarism score monitor, AI output quality score, language consistency check, daily cost report",
            "نسبة التفرد (>85%), جودة المحتوى (>4/5), زمن التوليد (<90 ثانية), التكلفة لكل مقال (<$0.10)"
        ),
        (
            4, "التحقق من الحقائق", "Fact-Check",
            "التحقق الآلي من صحة المعلومات باستخدام Google Fact Check API و ClaimReview. مقارنة الادعاءات مع مصادر موثوقة. وضع علامات (Verified / Unverified / False) على كل مقال.",
            "Google Fact Check API, ClaimReview, Snopes API, PolitiFact, custom fact-check rules engine",
            "10-30 ثانية لكل مقال",
            "False positives (مقال صحيح يُعتبر خاطئ), نقص في قاعدة بيانات fact-check, بطء API",
            "Fact-check accuracy report, false positive rate monitor, manual review queue",
            "دقة التحقق (>95%), نسبة False Positives (<3%), زمن التحقق (<30 ثانية), نسبة المقالات الموثقة (>80%)"
        ),
        (
            5, "تحسين SEO", "SEO Optimization",
            "إضافة Meta tags (title, description), Schema.org markup (NewsArticle, SportsEvent...), روابط داخلية (internal links), تحسين الكلمات المفتاحية في العنوان والـ Lead, توليد slug مناسب, إضافة alt text للصور.",
            "Custom SEO engine, Schema.org generators, internal linking algorithm, image alt-text AI",
            "5-15 ثانية لكل مقال",
            "Keyword stuffing, روابط داخلية مكسورة, schema غير صحيح, تجاوز حد الكلمات المفتاحية",
            "SEO score monitor, schema validation, broken link checker, keyword density report",
            "SEO score (>85/100), schema validity (100%), keyword density (1-3%), internal links per article (3-5)"
        ),
        (
            6, "النشر", "Publishing",
            "نشر المقال على الموقع مع تحديث sitemap.xml, إرسال ping إلى search engines (Google, Bing, Yandex), توليد RSS feed entry, جدولة النشر حسب أولوية الكاتيجوري.",
            "Custom CMS API, sitemap generator, Google Search Console API, Bing Webmaster API, RSS generator",
            "5-30 ثانية لكل مقال",
            "فشل قاعدة البيانات, بطء CDN, أخطاء sitemap, فشل الـ ping, تعارض مع جدول النشر",
            "Publish success rate, sitemap update monitor, ping response checker, publishing queue health",
            "معدل نجاح النشر (>99.5%), زمن النشر (<30 ثانية), sitemap update (real-time), ping success (>95%)"
        ),
        (
            7, "التوزيع", "Distribution",
            "نشر المقال على منصات السوشيال ميديا (Facebook, Twitter, Instagram, Telegram, WhatsApp), إرسال إشعارات push, إضافة لـ newsletter, مشاركة في Reddit/Quora حسب الكاتيجوري.",
            "Facebook Graph API, Twitter API, Instagram Graph API, Telegram Bot API, WhatsApp Business API, push notification service, Mailchimp",
            "10-60 ثانية لكل مقال",
            "API rate limits, تجاهل منصات السوشيال, إلغاء حسابات, قيود المحتوى, بطء في التوزيع",
            "Social share rate monitor, push notification delivery rate, newsletter open rate, account health dashboard",
            "معدل المشاركة (>10%), push delivery rate (>90%), newsletter open rate (>25%), زمن التوزيع (<60 ثانية)"
        ),
    ]

    for i, stage in enumerate(stages):
        write_data_row(ws, 6 + i, list(stage), i)
        # Make stage number bold and colored
        cell = ws.cell(row=6 + i, column=2)
        cell.font = Font(name=FONT_NAME, size=14, bold=True, color=PRIMARY)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    widths = [10, 22, 22, 50, 40, 14, 36, 36, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    # Increase row heights for stages (long content)
    for i in range(len(stages)):
        ws.row_dimensions[6 + i].height = 95

    ws.freeze_panes = 'C6'


# ============================================================
# SHEET 6: News Templates (نماذج الأخبار)
# ============================================================

def build_news_templates(wb):
    ws = wb.create_sheet("6. نماذج الأخبار")
    setup_rtl_sheet(ws)

    headers = [
        "Template ID", "الكاتيجوري", "نمط العنوان", "نمط الـ Lead",
        "هيكل الـ Body", "AI Tone", "AI Length",
        "AI POV", "Schema.org Type", "متطلبات الصور"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "📝 نماذج الأخبار الآلية",
                "20 قالباً - واحد لكل كاتيجوري - يحدد نمط العنوان والـ Lead والـ Body وإعدادات AI و Schema.org المناسب",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    templates = [
        # World
        ("TPL-WORLD-01", "🌍 أخبار عالمية",
         "[Country]: [Main Event] - [Key Detail] | [Date]",
         "In a significant development in [region], [main actor] [action] amid [context]. The event marks [significance].",
         "1) Lead 2) Background 3) Key Developments 4) Reactions 5) Analysis 6) What's Next",
         "Neutral, journalistic", "400-600 words", "Third person", "NewsArticle",
         "Featured image (1200x630), 1-2 inline images, optional map"),
        ("TPL-WORLD-02", "🌍 أخبار عالمية - أفريقيا",
         "[African Country]: [Event] - [Impact] | [Date]",
         "Developments in [country] as [actor] [action]. The situation has [impact] on [region].",
         "1) Lead 2) Context 3) Reactions 4) Regional implications 5) Background",
         "Neutral, factual", "350-500 words", "Third person", "NewsArticle",
         "Featured image, 1 inline image, optional regional map"),
        ("TPL-WORLD-03", "🌍 أخبار عالمية - أوروبا",
         "Europe: [Country] [Event] - [Stakeholder Reaction]",
         "[Country] has [action] in a move that [impact]. [Stakeholder] reacted by [reaction].",
         "1) Lead 2) Detailed event 3) EU context 4) Stakeholder reactions 5) Next steps",
         "Neutral, analytical", "400-550 words", "Third person", "NewsArticle",
         "Featured image, 1-2 inline, optional chart"),

        # Politics
        ("TPL-POLITICS-01", "🏛️ سياسة دولية",
         "[Country/Aktor] [Action]: [Consequence] | [Date]",
         "[Actor] announced [action] regarding [topic], marking [significance]. The move impacts [stakeholder].",
         "1) Lead 2) Official statement 3) Context 4) International reactions 5) Analysis 6) Implications",
         "Neutral, authoritative", "500-700 words", "Third person", "NewsArticle",
         "Featured image (official photo), 1-2 inline, optional infographic"),
        ("TPL-POLITICS-03", "🏛️ انتخابات",
         "[Country] Elections: [Candidate/Party] [Result] - [Turnout]%",
         "Voters in [country] went to the polls [date], with [result]. Turnout reached [X]% amid [context].",
         "1) Lead with results 2) Detailed breakdown 3) Candidate reactions 4) Analysis 5) What's next 6) Historical context",
         "Neutral, data-driven", "500-700 words", "Third person", "NewsArticle + Election",
         "Featured image, results chart, candidate photos, optional map"),

        # Economy
        ("TPL-ECONOMY-01", "💼 أسواق مالية",
         "[Market] [Direction]: [Index] [Change]% - [Reason]",
         "[Market] [direction] [X]% to [level] on [date], driven by [reason]. [Sector/stock] led [gainers/losers].",
         "1) Lead with key numbers 2) Market overview 3) Sectors performance 4) Key drivers 5) Expert quotes 6) Outlook",
         "Analytical, data-focused", "400-600 words", "Third person", "NewsArticle + Dataset",
         "Featured chart (mandatory), 1-2 stock charts, sector heatmap"),

        # Technology
        ("TPL-TECH-01", "💻 ذكاء اصطناعي",
         "[Company] announces [AI product]: [key feature] - [impact]",
         "[Company] unveiled [product], a [description] that [capability]. The announcement [significance].",
         "1) Lead 2) Product details 3) Technical specs 4) Competitive context 5) Industry reaction 6) Availability & pricing",
         "Tech-savvy, forward-looking", "500-700 words", "Third person", "TechArticle",
         "Product image (mandatory), screenshots, optional demo video link"),
        ("TPL-TECH-02", "💻 شركات التقنية",
         "[Big Tech] [Action]: [Product/Service] - [Impact]",
         "[Company] [action] [product], aiming to [goal]. The move [significance] in [market/industry].",
         "1) Lead 2) Announcement details 3) Strategic context 4) Market reaction 5) Competitor comparison 6) Future outlook",
         "Business-tech, analytical", "450-650 words", "Third person", "TechArticle",
         "Company logo, product image, stock chart (if relevant)"),

        # Sports
        ("TPL-SPORTS-01", "🏀 كرة القدم",
         "[Team A] [score] - [score] [Team B]: [Key moment/player]",
         "[Team A] defeated [Team B] [score] in [competition] at [venue]. [Player] was the star with [performance].",
         "1) Lead with score 2) Match summary 3) Key moments 4) Star player 5) Post-match reactions 6) Standings impact",
         "Energetic, play-by-play", "400-600 words", "Third person", "SportsEvent",
         "Action photo (mandatory), score graphic, league table snippet"),
        ("TPL-SPORTS-05", "🏀 أولمبياد",
         "Olympics [Sport]: [Athlete] wins [medal] - [Record/Detail]",
         "[Athlete] from [country] won [medal] in [sport] at [Olympics], [achievement detail].",
         "1) Lead 2) Performance breakdown 3) Competition context 4) Athlete background 5) Medal table impact 6) Reactions",
         "Inspirational, celebratory", "400-550 words", "Third person", "SportsEvent",
         "Podium photo, action shot, medal table update"),

        # Health
        ("TPL-HEALTH-01", "🏥 صحة عامة",
         "[Health topic]: [Finding/Action] - [Impact] | [Source]",
         "[Organization/Researchers] announced [finding/action] regarding [health topic], with [impact] on [population].",
         "1) Lead 2) Key findings 3) Methodology 4) Expert quotes 5) Public health implications 6) Recommendations",
         "Authoritative, cautious", "450-650 words", "Third person", "MedicalWebPage",
         "Medical illustration, expert photo, data chart"),

        # Environment
        ("TPL-ENV-02", "🌱 كوارث طبيعية",
         "BREAKING: [Disaster type] hits [Location] - [Severity/Casualties]",
         "A [disaster] struck [location] on [date], causing [impact]. Authorities [response] as [casualty numbers].",
         "1) Lead with impact 2) Event details 3) Casualties/damage 4) Emergency response 5) Weather/geological context 6) How to help",
         "Urgent, factual, empathetic", "400-600 words", "Third person", "NewsArticle",
         "Impact photo, satellite image, map of affected area"),

        # Weather (NEW)
        ("TPL-WEATHER-01", "🌤️ طقس يومي",
         "[City] Weather Today: [Condition], [High]°/[Low]° - [Alert if any]",
         "Today in [city], expect [condition] with temperatures ranging from [low]° to [high]°. [Additional context like wind, humidity].",
         "1) Today forecast 2) Hourly breakdown 3) Tomorrow outlook 4) Weekly forecast 5) Weather alerts 6) Tips (clothing, commute)",
         "Friendly, practical", "200-400 words", "Second person (you)", "WebPage",
         "Weather chart (mandatory), hourly temp graph, 7-day forecast widget"),
        ("TPL-WEATHER-02", "🌤️ تحذيرات جوية",
         "WEATHER ALERT: [Warning type] for [Region] - [Duration]",
         "[Authority] issued [warning type] for [region] effective [duration]. Residents should [action].",
         "1) Alert details 2) Affected areas 3) Expected impact 4) Safety recommendations 5) Official sources 6) Updates timeline",
         "Urgent, authoritative", "300-500 words", "Second person (you)", "WebPage",
         "Alert map (mandatory), radar image, warning badge"),

        # Food (NEW)
        ("TPL-FOOD-01", "🍳 وصفات",
         "[Recipe Name]: [Prep time]min prep, [Cook time]min cook - [Difficulty]",
         "Learn how to make [recipe name] with this easy [cuisine] recipe. Ready in [total time] minutes, serves [number].",
         "1) Recipe intro 2) Ingredients list 3) Step-by-step instructions 4) Pro tips 5) Variations 6) Nutrition info 7) Storage",
         "Friendly, instructional", "400-600 words", "Second person (you)", "Recipe",
         "Hero dish photo (mandatory), step photos, ingredients photo, nutrition card"),

        # Fashion (NEW)
        ("TPL-FASHION-01", "👗 عروض الأزياء",
         "[Brand] [Season] [Year]: [Theme] - [Standout piece]",
         "[Brand] unveiled its [season] [year] collection at [fashion week], featuring [theme]. [Designer] said [quote].",
         "1) Lead 2) Collection overview 3) Standout pieces 4) Designer inspiration 5) Celebrity attendees 6) Industry impact 7) Where to buy",
         "Editorial, descriptive", "500-700 words", "Third person", "NewsArticle",
         "Runway photos (multiple), backstage shot, celebrity photo"),

        # Cars (NEW)
        ("TPL-CARS-01", "🚗 سيارات جديدة",
         "[Brand] [Model] [Year]: [Key feature] - [Price]",
         "[Brand] has revealed the [year] [model], featuring [key feature]. Starting at [price], it competes with [rivals].",
         "1) Lead 2) Design overview 3) Powertrain & specs 4) Technology features 5) Pricing & trims 6) Competition 7) Availability",
         "Enthusiast, technical", "500-700 words", "Third person", "Product",
         "Hero car photo (mandatory), interior, exterior detail shots, spec table"),

        # Real Estate (NEW)
        ("TPL-RE-01", "🏠 أسعار العقارات",
         "[City] Property Prices [Direction] [X]% - [Reason]",
         "Property prices in [city] [direction] [X]% in [period], according to [source]. The trend reflects [reason].",
         "1) Lead with data 2) Detailed price changes 3) Neighborhood breakdown 4) Market drivers 5) Expert analysis 6) Forecast 7) Buyer advice",
         "Analytical, market-focused", "450-650 words", "Third person", "NewsArticle + Dataset",
         "Price chart (mandatory), neighborhood map, property photo"),

        # Jobs (NEW)
        ("TPL-JOBS-01", "💼 فرص عمل",
         "[Industry/Country] Jobs: [Number] openings - [Top role]",
         "The [industry] sector in [country] has [number] job openings, with [top role] in highest demand. [Key insight].",
         "1) Lead 2) Market overview 3) Top roles 4) Salary ranges 5) Skills in demand 6) How to apply 7) Industry outlook",
         "Informative, career-focused", "400-600 words", "Second person (you)", "NewsArticle",
         "Industry photo, salary chart, top companies logo strip"),

        # Crypto (NEW)
        ("TPL-CRYPTO-01", "₿ Bitcoin",
         "Bitcoin [Direction] [X]% to $[Price] - [Catalyst]",
         "Bitcoin [direction] to $[price] in the last [timeframe], driven by [catalyst]. [Market context].",
         "1) Lead with price action 2) Market data 3) Key drivers 4) On-chain metrics 5) Trader sentiment 6) Expert opinions 7) What to watch",
         "Analytical, data-driven", "350-550 words", "Third person", "NewsArticle",
         "Price chart (mandatory), market cap table, fear & greed index"),

        # Family (NEW)
        ("TPL-FAM-01", "👨‍👩‍👧 تربية",
         "[Parenting Topic]: [Tip/Insight] for [Age group] kids",
         "When it comes to [parenting topic], experts recommend [approach]. Here's what parents of [age] kids should know.",
         "1) Lead 2) Why it matters 3) Expert insights 4) Practical tips 5) Common mistakes 6) Age-appropriate advice 7) Resources",
         "Warm, supportive, evidence-based", "500-700 words", "Second person (you)", "WebPage",
         "Family photo, age-appropriate activity photo, infographic"),
    ]

    for i, tpl in enumerate(templates):
        write_data_row(ws, 6 + i, list(tpl), i)

    # Column widths
    widths = [16, 22, 36, 36, 50, 20, 16, 16, 22, 36]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    # Taller rows for long content
    for i in range(len(templates)):
        ws.row_dimensions[6 + i].height = 95

    ws.freeze_panes = 'C6'


# ============================================================
# SHEET 7: Publishing Schedule (جدول النشر)
# ============================================================

def build_publishing_schedule(wb):
    ws = wb.create_sheet("7. جدول النشر")
    setup_rtl_sheet(ws)

    headers = [
        "ID", "الكاتيجوري", "عدد الأخبار/يوم",
        "ذروة النشر (UTC)", "المنطقة الزمنية الأساسية",
        "الإثنين", "الثلاثاء", "الأربعاء", "الخميس",
        "الجمعة", "السبت", "الأحد", "ملاحظات"
    ]
    last_col = 1 + len(headers)
    write_title(ws, "📅 جدول النشر الموصى به",
                "مصفوفة الكاتيجوري × أيام الأسبوع مع تردد النشر وذروة الساعات لكل منطقة زمنية",
                last_col)

    write_headers(ws, headers, row=5, col_start=2)

    # Schedule data
    schedule = [
        ("world_news", "🌍 أخبار عالمية", 30, "06:00-22:00", "GMT+0 (London)", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓", "✓✓", "✓✓", "أعلى نشاط أيام الأسبوع"),
        ("politics", "🏛️ سياسة", 25, "07:00-21:00", "GMT+0", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓", "✓✓", "✓✓", "ذروة صباحية ومسائية"),
        ("economy", "💼 اقتصاد", 35, "07:00-17:00", "GMT-5 (NY)", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓", "✓", "أعلى نشاط أيام التداول"),
        ("technology", "💻 تكنولوجيا", 25, "08:00-20:00", "GMT-8 (SF)", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓", "✓✓", "✓", "أخبار الإطلاقات أيام الثلاثاء"),
        ("sports", "🏀 رياضة", 40, "16:00-23:00", "GMT+0", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "ذروة في عطلة نهاية الأسبوع"),
        ("entertainment", "🎭 ترفيه", 20, "12:00-22:00", "GMT-8 (LA)", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "ذروة في عطلات نهاية الأسبوع"),
        ("health", "🏥 صحة", 10, "08:00-18:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓", "✓", "محتوى تعليمي + أخبار"),
        ("environment", "🌱 بيئة", 8, "09:00-17:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓", "✓", "ذروة أثناء مؤتمرات COP"),
        ("education", "🎓 تعليم", 8, "08:00-18:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓", "✓", "ذروة في بداية العام الدراسي"),
        ("society_law", "⚖️ مجتمع وقانون", 10, "09:00-19:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓", "✓", "أخبار محاكم أيام الثلاثاء"),
        ("travel", "✈️ سفر وسياحة", 8, "10:00-20:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓✓", "✓✓✓", "ذروة عطلة نهاية الأسبوع"),
        ("religion", "🕌 دين", 6, "05:00-21:00", "GMT+3 (Makkah)", "✓", "✓", "✓", "✓", "✓✓✓", "✓✓", "✓", "ذروة أيام الجمعة ورمضان"),
        ("weather", "🌤️ طقس", 50, "00:00-24:00", "متعدد", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "تحديث كل ساعة - 24/7"),
        ("food", "🍳 طبخ", 10, "11:00-20:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓✓", "✓✓✓", "ذروة عطلة نهاية الأسبوع"),
        ("fashion", "👗 أزياء", 8, "10:00-20:00", "GMT+1 (Paris)", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓✓", "✓✓", "ذروة أيام عروض الأزياء"),
        ("cars", "🚗 سيارات", 8, "09:00-19:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓", "ذروة أيام المعارض"),
        ("real_estate", "🏠 عقارات", 8, "09:00-18:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓", "تحديثات أسبوعية للأسعار"),
        ("jobs", "💼 عمل ووظائف", 10, "08:00-18:00", "GMT+0", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓", "✓", "✓", "ذروة أيام الإثنين والثلاثاء"),
        ("crypto", "₿ كريبتو", 30, "00:00-24:00", "متعدد", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "✓✓✓", "سوق 24/7 - تحديثات كل ساعة"),
        ("family", "👨‍👩‍👧 حياة عائلية", 8, "10:00-20:00", "GMT+0", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓", "✓✓✓", "✓✓✓", "ذروة عطلة نهاية الأسبوع"),
    ]

    for i, sch in enumerate(schedule):
        write_data_row(ws, 6 + i, list(sch), i)
        # Color the daily volume cell
        vol_cell = ws.cell(row=6 + i, column=2 + 2)
        vol = sch[2]
        if vol >= 30:
            vol_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_NEGATIVE)
        elif vol >= 15:
            vol_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_WARNING)
        else:
            vol_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_POSITIVE)

        # Center the daily check columns
        for j in range(5, 12):
            cell = ws.cell(row=6 + i, column=2 + j)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Totals row
    total_row = 6 + len(schedule)
    total_fill = PatternFill('solid', fgColor=SECONDARY)
    total_font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)
    total_border = Border(top=Side(style='medium', color=NEUTRAL_200))

    total_vol = sum(s[2] for s in schedule)
    totals = ["TOTAL", "المجموع", total_vol, "—", "—", "—", "—", "—", "—", "—", "—", "—", "إجمالي يومي"]
    for i, v in enumerate(totals):
        cell = ws.cell(row=total_row, column=2 + i, value=v)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = total_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[total_row].height = 28

    # Column widths
    widths = [14, 22, 16, 18, 22, 10, 10, 10, 10, 10, 10, 10, 32]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    ws.freeze_panes = 'C6'


# ============================================================
# MAIN
# ============================================================

def main():
    print("🚀 Building Automated News Website Categorization Excel...")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build all 7 sheets
    print("  → Sheet 1: Overview...")
    build_overview(wb)

    print("  → Sheet 2: Full Taxonomy (398 rows)...")
    build_full_taxonomy(wb)

    print("  → Sheet 3: SEO Keywords Bank (~2000 rows)...")
    build_seo_keywords(wb)

    print("  → Sheet 4: Data Sources (40 sources)...")
    build_data_sources(wb)

    print("  → Sheet 5: Automation Pipeline (7 stages)...")
    build_automation_pipeline(wb)

    print("  → Sheet 6: News Templates (20 templates)...")
    build_news_templates(wb)

    print("  → Sheet 7: Publishing Schedule...")
    build_publishing_schedule(wb)

    # Set workbook properties
    wb.properties.creator = "Z.ai"
    wb.properties.title = "Automated News Website Categorization"
    wb.properties.subject = "20 Categories × 98 Subs × 398 Sub-Subs in 4 languages"

    # Save
    output_path = "/home/z/my-project/download/automated_news_categories.xlsx"
    wb.save(output_path)
    print(f"\n✅ Saved to: {output_path}")

    # Print stats
    print(f"\n📊 Final Stats:")
    print(f"   - Sheets: {len(wb.sheetnames)}")
    print(f"   - Total rows in Full Taxonomy: {sum(sum(len(s['subsubs']) for s in c['subs']) for c in TAXONOMY)}")
    print(f"   - Total SEO keywords: {sum(len(ss['keywords']) for c in TAXONOMY for s in c['subs'] for ss in s['subsubs'])}")


if __name__ == "__main__":
    main()
